import tkinter as tk
from tkinter import messagebox
from tkinter import simpledialog
from fpdf import FPDF
import openpyxl
import qrcode
from PIL import Image, ImageTk
import mysql.connector

# Replace these with your database credentials
db_host = "localhost"
db_user = "root"
db_password = ""
db_name = "savetask"

# Establish a connection to the database
db = mysql.connector.connect(
    host=db_host,
    user=db_user,
    password=db_password,
    database=db_name
)

# Create a cursor object to execute SQL queries
cursor = db.cursor()

def add_task():
    task = entry.get()
    if task:
        listbox.insert(tk.END, task)
        entry.delete(0, tk.END)

        # Insert the task into the database
        insert_query = "INSERT INTO todo_list (task) VALUES (%s)"
        cursor.execute(insert_query, (task,))
        db.commit()  # Commit the transaction to save the data to the database

    else:
        messagebox.showwarning("Warning", "You must enter a task.")

def delete_task():
    try:
        selected_task = listbox.curselection()[0]
        task_to_delete = listbox.get(selected_task)

        # Delete the selected task from the database
        delete_query = "DELETE FROM todo_list WHERE task = %s"
        cursor.execute(delete_query, (task_to_delete,))
        db.commit()  # Commit the transaction to update the database

        listbox.delete(selected_task)
    except IndexError:
        messagebox.showwarning("Warning", "You must select a task to delete.")

def edit_task():
    try:
        selected_task = listbox.curselection()[0]
        task_to_edit = listbox.get(selected_task)
        edited_task = simpledialog.askstring("Edit Task", "Edit the selected task:", initialvalue=task_to_edit)
        if edited_task:
            listbox.delete(selected_task)
            listbox.insert(selected_task, edited_task)

            # Update the edited task in the database
            update_query = "UPDATE todo_list SET task = %s WHERE task = %s"
            cursor.execute(update_query, (edited_task, task_to_edit))
            db.commit()  # Commit the transaction to update the database

    except IndexError:
        messagebox.showwarning("Warning", "You must select a task to edit.")

def fetch_data_from_db():
    # Clear the listbox
    listbox.delete(0, tk.END)

    # Fetch all tasks from the database
    select_query = "SELECT task FROM todo_list"
    cursor.execute(select_query)

    for task in cursor.fetchall():
        listbox.insert(tk.END, task[0])

def export_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    for task in listbox.get(0, tk.END):
        pdf.cell(200, 10, txt=task, ln=True)
    
    pdf_filename = "todo_list.pdf"
    pdf.output(pdf_filename)
    print("Tasks exported to PDF successfully.")
    

def export_excel():
    excel_filename = "todo_list.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    for i, task in enumerate(listbox.get(0, tk.END), start=1):
        sheet.cell(row=i, column=1, value=task)
    
    workbook.save(excel_filename)
    print("Tasks exported to Excel successfully.")

def export_qr_code():
    combined_tasks = "\n".join(listbox.get(0, tk.END))
    
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(combined_tasks)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img_filename = "todo_list_qr.png"
    img.save(img_filename)
    img.show()
    print("QR code generated successfully.")

root = tk.Tk()
root.title("TO-DO List")

# Increase the size of the main window
root.geometry("400x400")

frame = tk.Frame(root)
frame.pack(pady=10)

entry = tk.Entry(frame, font=("Helvetica", 14))
entry.pack(side=tk.LEFT, padx=10)

add_button = tk.Button(frame, text="Add Task", font=("Helvetica", 12), command=add_task)
add_button.pack(side=tk.LEFT)

edit_button = tk.Button(frame, text="Edit Task", font=("Helvetica", 12), command=edit_task)
edit_button.pack(side=tk.LEFT)

delete_button = tk.Button(root, text="Delete Task", font=("Helvetica", 12), command=delete_task)
delete_button.pack(pady=10)

fetch_button = tk.Button(root, text="Fetch Data", font=("Helvetica", 12), command=fetch_data_from_db)
fetch_button.pack()

export_pdf_button = tk.Button(root, text="Export to PDF", font=("Helvetica", 12), command=export_pdf)
export_pdf_button.pack()

export_excel_button = tk.Button(root, text="Export to Excel", font=("Helvetica", 12), command=export_excel)
export_excel_button.pack()

export_qr_button = tk.Button(root, text="Export QR Code", font=("Helvetica", 12), command=export_qr_code)
export_qr_button.pack()

listbox = tk.Listbox(root, selectmode=tk.SINGLE, font=("Helvetica", 12))
listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

root.mainloop()
